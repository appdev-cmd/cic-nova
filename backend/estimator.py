import os
import re
import io
import pandas as pd
import openpyxl
import xml.etree.ElementTree as ET
from copy import copy
from openpyxl.utils.dataframe import dataframe_to_rows
from database import get_db_connection

def determine_door_type(code, name=""):
    code_upper = str(code).upper()
    name_upper = str(name).upper()
    
    # 1. Vách kính
    if 'VKT' in code_upper or 'VÁCH' in name_upper or 'VKT' in name_upper:
        return 'VÁCH KÍNH'
        
    # 2. Cửa sổ
    if 'SWA' in code_upper or 'WA' in code_upper or 'CS' in code_upper or 'SỔ' in name_upper:
        return 'CỬA SỔ'
        
    # 3. Cửa đi
    if 'SDA' in code_upper or 'DA' in code_upper or 'CD' in code_upper or 'ĐI' in name_upper or 'CỬA ĐI' in name_upper:
        return 'CỬA ĐI'
        
    # 4. Fallback
    if 'WINDOW' in name_upper:
        return 'CỬA SỔ'
    if 'DOOR' in name_upper:
        return 'CỬA ĐI'
        
    return 'CỬA SỔ'

def parse_opera_file(project_id, file_path):
    """
    Parse Opera export (.xls, .xlsx, or .xml) and:
    1. Save material prices to project_material_prices.
    2. Automatically extract doors and create them in project_doors.
    3. Generate static formulas (profile & accessory formulas) for templates.
    """
    print(f"Parsing Opera file: {file_path} for Project: {project_id}")
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    file_ext = os.path.splitext(file_path)[1].lower()
    
    if file_ext == '.xml':
        imported_count = parse_opera_xml(project_id, file_path, cursor, conn)
    else:
        imported_count = parse_opera_excel(project_id, file_path, cursor, conn)
        
    conn.commit()
    conn.close()
    return imported_count

def parse_opera_xml(project_id, file_path, cursor, conn):
    print("Parsing Opera XML file...")
    tree = ET.parse(file_path)
    root = tree.getroot()
    
    # 1. Import material prices globally and project-specifically
    materials_inserted = 0
    unique_materials = {}
    
    for mat in root.findall(".//material"):
        code = (mat.findtext("mat_alternative_code") or mat.findtext("mat_supplier_code") or "").strip()
        name = (mat.findtext("mat_name") or mat.findtext("mat_description") or "").strip()
        unit = (mat.findtext("mat_unit") or "pc").strip()
        
        price_val = mat.findtext("mat_price")
        if not price_val:
            price_val = mat.findtext("mat_full_price")
        try:
            price = float(price_val) if price_val else 0.0
        except (ValueError, TypeError):
            price = 0.0
            
        weight_val = mat.findtext("mat_unit_weight") or mat.findtext("mat_weight")
        try:
            weight = float(weight_val) if weight_val else 0.0
        except (ValueError, TypeError):
            weight = 0.0
            
        if code and code != 'nan':
            unique_materials[code] = {
                'name': name,
                'unit': unit,
                'price': price,
                'weight': weight
            }
            
    for code, m in unique_materials.items():
        mat_category = 'other'
        if 'nhôm' in m['name'].lower() or 'khung' in m['name'].lower() or 'cánh' in m['name'].lower():
            mat_category = 'aluminum'
        elif 'kính' in m['name'].lower() or 'glass' in m['name'].lower():
            mat_category = 'glass'
        elif 'gioăng' in m['name'].lower() or 'keo' in m['name'].lower() or 'sealant' in m['name'].lower():
            mat_category = 'other'
        elif m['unit'] == 'pc':
            mat_category = 'accessory'
            
        cursor.execute("""
        INSERT INTO materials (code, name, category, unit, default_price, weight_per_m)
        VALUES (%s, %s, %s, %s, %s, %s)
        ON CONFLICT (code) DO UPDATE SET
            name = EXCLUDED.name,
            default_price = EXCLUDED.default_price,
            weight_per_m = EXCLUDED.weight_per_m
        """, (code, m['name'], mat_category, m['unit'], m['price'], m['weight']))
        
        cursor.execute("""
        INSERT INTO project_material_prices (project_id, material_code, material_name, unit, price, weight)
        VALUES (%s, %s, %s, %s, %s, %s)
        ON CONFLICT (project_id, material_code) DO UPDATE SET
            material_name = EXCLUDED.material_name,
            unit = EXCLUDED.unit,
            price = EXCLUDED.price,
            weight = EXCLUDED.weight
        """, (project_id, code, m['name'], m['unit'], m['price'], m['weight']))
        materials_inserted += 1

    # 2. Extract systems, templates, formulas and create doors
    components = root.findall(".//component")
    for comp in components:
        cmp_pos = (comp.findtext("cmp_position") or "").strip()
        cmp_name = (comp.findtext("cmp_name") or "").strip()
        cmp_system = (comp.findtext("cmp_system") or "Nova System").strip()
        
        try:
            width = float(comp.findtext("cmp_width") or 1000.0)
            height = float(comp.findtext("cmp_height") or 2000.0)
            qty = int(comp.findtext("cmp_quantity") or 1)
        except (ValueError, TypeError):
            width = 1000.0
            height = 2000.0
            qty = 1
            
        if not cmp_name:
            continue
            
        cursor.execute("SELECT id FROM systems WHERE name = %s", (cmp_system,))
        sys_row = cursor.fetchone()
        if sys_row:
            system_id = sys_row[0]
        else:
            cursor.execute("INSERT INTO systems (name) VALUES (%s) RETURNING id", (cmp_system,))
            system_id = cursor.fetchone()[0]
            
        cursor.execute("SELECT id FROM templates WHERE code = %s", (cmp_name,))
        tmpl_row = cursor.fetchone()
        
        door_type = determine_door_type(cmp_name, cmp_name)
        
        if tmpl_row:
            template_id = tmpl_row[0]
        else:
            cursor.execute("""
            INSERT INTO templates (system_id, code, name, type, accessory_brand, glass_type)
            VALUES (%s, %s, %s, %s, 'Draho', 'k8cl')
            RETURNING id
            """, (system_id, cmp_name, cmp_name, door_type))
            template_id = cursor.fetchone()[0]
            
        cursor.execute("DELETE FROM profile_formulas WHERE template_id = %s", (template_id,))
        cursor.execute("DELETE FROM accessory_formulas WHERE template_id = %s", (template_id,))
        
        comp_materials = comp.findall(".//materials/material")
        for mat in comp_materials:
            m_code = (mat.findtext("mat_alternative_code") or mat.findtext("mat_supplier_code") or "").strip()
            m_name = (mat.findtext("mat_name") or mat.findtext("mat_description") or "").strip()
            m_type = (mat.findtext("mat_type") or "").strip()
            
            try:
                m_qty = float(mat.findtext("mat_quantity") or 0.0)
                m_pieces = int(mat.findtext("mat_pieces") or 1)
            except (ValueError, TypeError):
                m_qty = 0.0
                m_pieces = 1
                
            try:
                m_unit_weight = float(mat.findtext("mat_unit_weight") or 0.0)
            except (ValueError, TypeError):
                m_unit_weight = 0.0
                
            if not m_code:
                continue
                
            qty_per_set = m_qty / qty if qty > 0 else m_qty
            
            if 'profile' in m_type.lower():
                cursor.execute("""
                INSERT INTO profile_formulas (template_id, name, code, dimension_type, formula, qty, weight_per_m)
                VALUES (%s, %s, %s, 'W', %s, %s, %s)
                """, (template_id, m_name, m_code, f"{qty_per_set:.4f}", m_pieces, m_unit_weight))
            else:
                cursor.execute("""
                INSERT INTO accessory_formulas (template_id, name, code, qty)
                VALUES (%s, %s, %s, %s)
                """, (template_id, m_name, m_code, qty_per_set))
                
        panes = comp.findall(".//cmp_panes_details/pane_details")
        for idx, pane in enumerate(panes):
            g_code = (pane.findtext("mat_name") or f"glass_{template_id}").strip()
            g_name = (pane.findtext("mat_description") or "Kính").strip()
            try:
                g_w = float(pane.findtext("mat_width") or 0.0)
                g_h = float(pane.findtext("mat_height") or 0.0)
                g_qty = float(pane.findtext("mat_quantity") or 1.0)
                area = (g_w * g_h / 1000000.0) * g_qty
            except (ValueError, TypeError):
                area = 0.0
            
            qty_per_set = area / qty if qty > 0 else area
            if qty_per_set > 0:
                cursor.execute("""
                INSERT INTO accessory_formulas (template_id, name, code, qty)
                VALUES (%s, %s, %s, %s)
                """, (template_id, g_name, g_code, qty_per_set))
                
        cmp_desc = (comp.findtext("cmp_description") or comp.findtext("cmp_notes") or "").strip()
        cursor.execute("SELECT id FROM project_doors WHERE project_id = %s AND code = %s", (project_id, cmp_pos))
        door_row = cursor.fetchone()
        if door_row:
            cursor.execute("""
            UPDATE project_doors 
            SET template_id = %s, width = %s, height = %s, qty = %s,
                description = COALESCE(NULLIF(%s, ''), description)
            WHERE id = %s
            """, (template_id, width, height, qty, cmp_desc, door_row[0]))
        else:
            cursor.execute("""
            INSERT INTO project_doors (project_id, code, template_id, width, height, qty, description)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
            """, (project_id, cmp_pos, template_id, width, height, qty, cmp_desc))
            
    return materials_inserted

def parse_opera_excel(project_id, file_path, cursor, conn):
    print("Parsing Opera Excel file...")
    df = pd.read_excel(file_path, engine='xlrd')
    df.columns = [col.strip() if isinstance(col, str) else col for col in df.columns]
    
    if 'Typology Name' in df.columns:
        return parse_typologies_excel_format(project_id, df, cursor, conn)
        
    imported_count = 0
    for idx, row in df.iterrows():
        code = str(row.get('Code', '')).strip()
        name = str(row.get('Description', '')).strip()
        if not name or name == 'nan':
            name = str(row.get('Name', '')).strip()
            
        unit = str(row.get('QuantityUnit', 'pc')).strip()
        price_val = row.get('ID')
        if pd.isna(price_val) or price_val == '':
            price_val = row.get('Unit Price')
            
        unit_weight = row.get('Unit Weight')
        
        try:
            price = float(price_val) if not pd.isna(price_val) else 0.0
        except (ValueError, TypeError):
            price = 0.0
            
        try:
            weight = float(unit_weight) if not pd.isna(unit_weight) else 0.0
        except (ValueError, TypeError):
            weight = 0.0
            
        if not code or code == 'nan':
            continue
            
        cursor.execute("""
        INSERT INTO project_material_prices (project_id, material_code, material_name, unit, price, weight)
        VALUES (%s, %s, %s, %s, %s, %s)
        ON CONFLICT(project_id, material_code) DO UPDATE SET
            material_name=excluded.material_name,
            unit=excluded.unit,
            price=excluded.price,
            weight=excluded.weight
        """, (project_id, code, name, unit, price, weight))
        imported_count += 1
        
    return imported_count

def parse_typologies_excel_format(project_id, df, cursor, conn):
    print("Detected DINH MUC TYPOLOGIES Excel format...")
    unique_materials = {}
    for idx, row in df.iterrows():
        code = str(row.get('Code', '')).strip()
        name = str(row.get('Description', '')).strip()
        if not name or name == 'nan':
            name = str(row.get('Name', '')).strip()
        unit = str(row.get('QuantityUnit', 'pc')).strip()
        
        price_val = row.get('Unit Price')
        try:
            price = float(price_val) if not pd.isna(price_val) else 0.0
        except (ValueError, TypeError):
            price = 0.0
            
        weight_val = row.get('Unit Weight')
        try:
            weight = float(weight_val) if not pd.isna(weight_val) else 0.0
        except (ValueError, TypeError):
            weight = 0.0
            
        if code and code != 'nan':
            unique_materials[code] = {
                'name': name,
                'unit': unit,
                'price': price,
                'weight': weight
            }
            
    materials_inserted = 0
    for code, m in unique_materials.items():
        mat_category = 'other'
        if 'nhôm' in m['name'].lower() or 'khung' in m['name'].lower() or 'cánh' in m['name'].lower():
            mat_category = 'aluminum'
        elif 'kính' in m['name'].lower() or 'glass' in m['name'].lower():
            mat_category = 'glass'
        elif 'gioăng' in m['name'].lower() or 'keo' in m['name'].lower() or 'sealant' in m['name'].lower():
            mat_category = 'other'
        elif m['unit'] == 'pc':
            mat_category = 'accessory'
            
        cursor.execute("""
        INSERT INTO materials (code, name, category, unit, default_price, weight_per_m)
        VALUES (%s, %s, %s, %s, %s, %s)
        ON CONFLICT (code) DO UPDATE SET
            name = EXCLUDED.name,
            default_price = EXCLUDED.default_price,
            weight_per_m = EXCLUDED.weight_per_m
        """, (code, m['name'], mat_category, m['unit'], m['price'], m['weight']))
        
        cursor.execute("""
        INSERT INTO project_material_prices (project_id, material_code, material_name, unit, price, weight)
        VALUES (%s, %s, %s, %s, %s, %s)
        ON CONFLICT (project_id, material_code) DO UPDATE SET
            material_name = EXCLUDED.material_name,
            unit = EXCLUDED.unit,
            price = EXCLUDED.price,
            weight = EXCLUDED.weight
        """, (project_id, code, m['name'], m['unit'], m['price'], m['weight']))
        materials_inserted += 1
        
    system_name = "Nova Excel System"
    cursor.execute("SELECT id FROM systems WHERE name = %s", (system_name,))
    sys_row = cursor.fetchone()
    if sys_row:
        system_id = sys_row[0]
    else:
        cursor.execute("INSERT INTO systems (name) VALUES (%s) RETURNING id", (system_name,))
        system_id = cursor.fetchone()[0]
        
    grouped = df.groupby('Typology Name')
    for typo_name, group in grouped:
        if not isinstance(typo_name, str) or not typo_name.strip():
            continue
            
        typo_name = typo_name.strip()
        
        cursor.execute("SELECT id FROM templates WHERE code = %s", (typo_name,))
        tmpl_row = cursor.fetchone()
        
        door_type = determine_door_type(typo_name, typo_name)
        
        if tmpl_row:
            template_id = tmpl_row[0]
        else:
            cursor.execute("""
            INSERT INTO templates (system_id, code, name, type, accessory_brand, glass_type)
            VALUES (%s, %s, %s, %s, 'Draho', 'k8cl')
            RETURNING id
            """, (system_id, typo_name, typo_name, door_type))
            template_id = cursor.fetchone()[0]
            
        cursor.execute("DELETE FROM profile_formulas WHERE template_id = %s", (template_id,))
        cursor.execute("DELETE FROM accessory_formulas WHERE template_id = %s", (template_id,))
        
        door_w, door_h = 1000.0, 2000.0
        glass_rows = group[group['QuantityUnit'] == 'm2']
        if not glass_rows.empty:
            max_glass = glass_rows.loc[glass_rows['Quantity'].idxmax()]
            try:
                g_w = float(max_glass.get('Width', 0.0))
                g_h = float(max_glass.get('Height', 0.0))
                if g_w > 0: door_w = g_w + 100
                if g_h > 0: door_h = g_h + 100
            except:
                pass
                
        for idx, row in group.iterrows():
            m_code = str(row.get('Code', '')).strip()
            m_name = str(row.get('Description', '')).strip()
            if not m_name or m_name == 'nan':
                m_name = str(row.get('Name', '')).strip()
            unit = str(row.get('QuantityUnit', 'pc')).strip()
            
            try:
                quantity = float(row.get('Quantity', 0.0))
                pieces = int(row.get('Pieces', 1))
            except:
                quantity = 0.0
                pieces = 1
                
            try:
                unit_weight = float(row.get('Unit Weight', 0.0))
            except:
                unit_weight = 0.0
                
            if not m_code or m_code == 'nan':
                continue
                
            qty_per_set = quantity
            
            if unit == 'm' and unit_weight > 0:
                cursor.execute("""
                INSERT INTO profile_formulas (template_id, name, code, dimension_type, formula, qty, weight_per_m)
                VALUES (%s, %s, %s, 'W', %s, %s, %s)
                """, (template_id, m_name, m_code, f"{qty_per_set:.4f}", pieces, unit_weight))
            else:
                cursor.execute("""
                INSERT INTO accessory_formulas (template_id, name, code, qty)
                VALUES (%s, %s, %s, %s)
                """, (template_id, m_name, m_code, qty_per_set))
                
        # Extract custom description from the typology group if present
        typo_desc = None
        for col_name in ['typology description', 'description', 'comments', 'notes', 'ghi chú', 'mô tả']:
            matched_col = next((c for c in group.columns if str(c).lower().strip() == col_name), None)
            if matched_col:
                non_empty = group[group[matched_col].notna()][matched_col]
                if not non_empty.empty:
                    val = str(non_empty.iloc[0]).strip()
                    if val and val.lower() != 'nan':
                        typo_desc = val
                        break

        cursor.execute("SELECT id FROM project_doors WHERE project_id = %s AND code = %s", (project_id, typo_name))
        door_row = cursor.fetchone()
        if door_row:
            cursor.execute("""
            UPDATE project_doors 
            SET template_id = %s, width = %s, height = %s, qty = 1,
                description = COALESCE(NULLIF(%s, ''), description)
            WHERE id = %s
            """, (template_id, door_w, door_h, typo_desc, door_row[0]))
        else:
            cursor.execute("""
            INSERT INTO project_doors (project_id, code, template_id, width, height, qty, description)
            VALUES (%s, %s, %s, %s, %s, 1, %s)
            """, (project_id, typo_name, template_id, door_w, door_h, typo_desc))
            
    return materials_inserted

def eval_formula(formula_str, variables):
    """
    Safely evaluate formula string using variables dict.
    Variables include: W, H, W1, H1, W2, H2
    """
    # Clean formula string: remove spaces, convert to lowercase for comparison
    clean_formula = formula_str.upper().strip()
    
    # Replace variable names with their values
    for var, val in variables.items():
        if val is None:
            val = 0.0
        # Replace using regex to match word boundaries so 'W' doesn't replace 'W1'
        clean_formula = re.sub(r'\b' + var + r'\b', str(val), clean_formula)
        
    # Standardize basic safety check: only allow numbers, math operators
    if not re.match(r'^[\d.+\-*/()\s]+$', clean_formula):
        # If it has other characters, check if it's just a number
        try:
            return float(clean_formula)
        except ValueError:
            print(f"Warning: Invalid characters in formula evaluation: {formula_str} -> {clean_formula}")
            return 0.0
            
    try:
        result = eval(clean_formula, {"__builtins__": None}, {})
        return float(result)
    except Exception as e:
        print(f"Error evaluating formula '{formula_str}' (processed: '{clean_formula}'): {e}")
        return 0.0

def calculate_project_estimates(project_id):
    """
    Perform full project estimation calculations.
    Returns detail data of each door item.
    """
    conn = get_db_connection()
    cursor = conn.cursor()
    
    cursor.execute("SELECT has_opera_bom, price_book_id, target_profit_margin, target_total_price FROM projects WHERE id = %s", (project_id,))
    proj_row = cursor.fetchone()
    has_opera_bom = proj_row['has_opera_bom'] if proj_row and 'has_opera_bom' in proj_row else False
    price_book_id = proj_row['price_book_id'] if proj_row else None
    target_profit_margin = proj_row['target_profit_margin'] if proj_row and proj_row['target_profit_margin'] is not None else 10.0
    target_total_price = proj_row['target_total_price'] if proj_row and proj_row['target_total_price'] is not None else 0.0

    if has_opera_bom:
        cursor.execute("""
        SELECT pd.*, pd.code as template_code, pd.template_name, '' as door_type,
               0.0 as percent_aluminum, 0.0 as percent_glass, 0.0 as percent_accessories, 0.0 as percent_labor,
               'Opera Glass' as glass_type, '' as accessory_brand, '' as system_name
        FROM project_doors pd
        WHERE pd.project_id = %s
        """, (project_id,))
        doors = [dict(row) for row in cursor.fetchall()]
        for d in doors:
            d['door_type'] = determine_door_type(d['code'], d['template_name'])
    else:
        cursor.execute("""
        SELECT pd.*, t.code as template_code, t.name as template_name, t.type as door_type,
               t.percent_aluminum, t.percent_glass, t.percent_accessories, t.percent_labor,
               t.glass_type, t.accessory_brand, s.name as system_name
        FROM project_doors pd
        JOIN templates t ON pd.template_id = t.id
        LEFT JOIN systems s ON t.system_id = s.id
        WHERE pd.project_id = %s
        """, (project_id,))
        doors = [dict(row) for row in cursor.fetchall()]

    cursor.execute("""
    SELECT pics.cost_type, pics.custom_value, icc.option_name, icc.value_type, icc.value
    FROM project_indirect_cost_selections pics
    LEFT JOIN indirect_cost_configs icc ON pics.indirect_cost_config_id = icc.id
    WHERE pics.project_id = %s
    """, (project_id,))
    indirect_selections = {row['cost_type']: {
        'custom_value': row['custom_value'],
        'option_name': row['option_name'],
        'value_type': row['value_type'],
        'value': row['value']
    } for row in cursor.fetchall()}

    project_total_area = 0.0
    for door in doors:
        w = door['width']
        h = door['height']
        w_m = w if w < 10.0 else w / 1000.0
        h_m = h if h < 10.0 else h / 1000.0
        project_total_area += (w_m * h_m) * door['qty']

    trans_sel = indirect_selections.get('transport')
    project_transport_total = 0.0
    if trans_sel:
        if trans_sel['custom_value'] is not None:
            project_transport_total = trans_sel['custom_value']
        elif trans_sel['value'] is not None and trans_sel['value_type'] == 'fixed':
            project_transport_total = trans_sel['value']

    cursor.execute("SELECT material_code, price, weight, material_name, unit FROM project_material_prices WHERE project_id = %s", (project_id,))
    prices = {row['material_code']: {
        'price': row['price'], 
        'weight': row['weight'], 
        'name': row['material_name'],
        'unit': row['unit']
    } for row in cursor.fetchall()}
    
    book_prices = {}
    if price_book_id:
        cursor.execute("SELECT material_code, price FROM material_price_book_items WHERE price_book_id = %s", (price_book_id,))
        book_prices = {row['material_code']: row['price'] for row in cursor.fetchall()}
    
    cursor.execute("SELECT code, default_price, weight_per_m, name, unit FROM materials")
    global_materials = {row['code']: {
        'price': row['default_price'],
        'weight': row['weight_per_m'],
        'name': row['name'],
        'unit': row['unit']
    } for row in cursor.fetchall()}
    
    materials_by_typo = {}
    if has_opera_bom:
        cursor.execute("""
        SELECT pom.*, m.category as catalog_category
        FROM project_opera_materials pom
        LEFT JOIN materials m ON pom.mapped_material_id = m.id
        WHERE pom.project_id = %s
        """, (project_id,))
        all_materials = cursor.fetchall()
        for mat in all_materials:
            typo = mat['typology_name']
            if typo not in materials_by_typo:
                materials_by_typo[typo] = []
            materials_by_typo[typo].append(mat)
            
    results = []
    for door in doors:
        door_id = door['id']
        template_id = door['template_id']
        door_code = door['code']
        door_name = door['template_name']
        door_type = door['door_type']
        
        w = door['width']
        h = door['height']
        w_m = w if w < 10.0 else w / 1000.0
        h_m = h if h < 10.0 else h / 1000.0
        
        w1 = door.get('width1')
        h1 = door.get('height1')
        w1_m = w1 if (w1 is not None and w1 < 10.0) else (w1 / 1000.0 if w1 is not None else None)
        h1_m = h1 if (h1 is not None and h1 < 10.0) else (h1 / 1000.0 if h1 is not None else None)
        w2 = door.get('width2')
        h2 = door.get('height2')
        w2_m = w2 if (w2 is not None and w2 < 10.0) else (w2 / 1000.0 if w2 is not None else None)
        h2_m = h2 if (h2 is not None and h2 < 10.0) else (h2 / 1000.0 if h2 is not None else None)
        
        qty_sets = door['qty']
        area = w_m * h_m
        total_area = area * qty_sets
        variables = {'W': w_m, 'H': h_m, 'W1': w1_m, 'H1': h1_m, 'W2': w2_m, 'H2': h2_m}
        
        if has_opera_bom:
            profiles_cost = 0.0
            profiles_weight = 0.0
            profiles_details = []
            acc_cost = 0.0
            acc_details = []
            glass_cost = 0.0
            glass_qty_total = 0.0
            other_cost = 0.0
            
            door_mats = materials_by_typo.get(door_code, [])
            for mat in door_mats:
                cat = mat['catalog_category']
                if not cat:
                    code_lower = mat['code'].lower()
                    name_lower = mat['name'].lower()
                    unit_lower = mat['quantity_unit'].lower()
                    if 'glass' in code_lower or 'kinh' in code_lower or 'kinh' in name_lower or 'gl' in code_lower:
                        cat = 'glass'
                    elif 'ac' in code_lower or 'pk' in code_lower or 'phu kien' in name_lower or 'accessory' in code_lower:
                        cat = 'accessory'
                    elif 'al' in code_lower or 'nhom' in name_lower or mat['unit_weight'] is not None or unit_lower in ['m', 'meter', 'kg']:
                        cat = 'aluminum'
                    else:
                        cat = 'other'
                
                qty = mat['quantity']
                unit_price = mat['unit_price'] if mat['unit_price'] is not None else 0.0
                cost = qty * unit_price
                
                if cat == 'aluminum':
                    profiles_cost += cost
                    weight_per_m = mat['unit_weight'] or 0.0
                    weight_total = qty * weight_per_m if mat['quantity_unit'].lower() in ['m', 'meter'] else (qty if mat['quantity_unit'].lower() == 'kg' else 0.0)
                    profiles_weight += weight_total
                    profiles_details.append({
                        'name': mat['name'],
                        'code': mat['code'],
                        'formula': 'Opera BOM',
                        'length': mat['width'] or qty,
                        'qty': 1,
                        'weight_per_m': weight_per_m,
                        'total_weight': weight_total,
                        'unit_price': unit_price,
                        'total_price': cost
                    })
                elif cat == 'accessory':
                    acc_cost += cost
                    acc_details.append({
                        'name': mat['name'],
                        'code': mat['code'],
                        'qty': qty,
                        'unit_price': unit_price,
                        'total_price': cost
                    })
                elif cat == 'glass':
                    glass_cost += cost
                    glass_qty_total += qty
                else:
                    other_cost += cost
            
            sum_materials = profiles_cost + glass_cost + acc_cost + other_cost
            glass_type = 'Opera Glass'
            glass_price = glass_cost / glass_qty_total if glass_qty_total > 0 else 0.0
            glass_area = glass_qty_total
        else:
            cursor.execute("SELECT * FROM profile_formulas WHERE template_id = %s", (template_id,))
            profile_formulas = cursor.fetchall()
            
            profiles_cost = 0.0
            profiles_weight = 0.0
            profiles_details = []
            for pf in profile_formulas:
                pf_code = pf['code']
                pf_name = pf['name']
                formula = pf['formula']
                pf_qty = pf['qty']
                if pf_code in prices:
                    unit_price = prices[pf_code]['price']
                    weight_per_m = prices[pf_code]['weight'] if prices[pf_code]['weight'] > 0 else pf['weight_per_m']
                elif pf_code in book_prices:
                    unit_price = book_prices[pf_code]
                    weight_per_m = global_materials.get(pf_code, {}).get('weight', pf['weight_per_m'])
                elif pf_code in global_materials:
                    unit_price = global_materials[pf_code]['price']
                    weight_per_m = global_materials[pf_code]['weight'] if global_materials[pf_code]['weight'] > 0 else pf['weight_per_m']
                else:
                    unit_price = 98000.0
                    weight_per_m = pf['weight_per_m']
                
                length = eval_formula(formula, variables)
                weight_total = pf_qty * length * weight_per_m
                cost = weight_total * unit_price
                profiles_cost += cost
                profiles_weight += weight_total
                profiles_details.append({'name': pf_name, 'code': pf_code, 'formula': formula, 'length': length, 'qty': pf_qty, 'weight_per_m': weight_per_m, 'total_weight': weight_total, 'unit_price': unit_price, 'total_price': cost})
                
            cursor.execute("SELECT * FROM accessory_formulas WHERE template_id = %s", (template_id,))
            acc_formulas = cursor.fetchall()
            acc_cost = 0.0
            acc_details = []
            for af in acc_formulas:
                af_code = af['code']
                af_name = af['name']
                af_qty = af['qty']
                if af_code in prices: unit_price = prices[af_code]['price']
                elif af_code in book_prices: unit_price = book_prices[af_code]
                elif af_code in global_materials: unit_price = global_materials[af_code]['price']
                else: unit_price = 0.0
                cost = af_qty * unit_price
                acc_cost += cost
                acc_details.append({'name': af_name, 'code': af_code, 'qty': af_qty, 'unit_price': unit_price, 'total_price': cost})
                
            glass_type = door['glass_type']
            glass_price = 0.0
            glass_code = None
            for code in prices:
                if 'glass' in code.lower() or 'kinh' in code.lower():
                    if glass_type.lower() in code.lower() or code.lower() in glass_type.lower():
                        glass_code = code
                        break
            if glass_code: glass_price = prices[glass_code]['price']
            else:
                for code in book_prices:
                    if 'glass' in code.lower() or 'kinh' in code.lower():
                        if glass_type.lower() in code.lower() or code.lower() in glass_type.lower():
                            glass_code = code
                            break
                if glass_code: glass_price = book_prices[glass_code]
                else:
                    for code in global_materials:
                        if 'glass' in code.lower() or 'kinh' in code.lower():
                            if glass_type.lower() in code.lower() or code.lower() in glass_type.lower():
                                glass_code = code
                                break
                    if glass_code: glass_price = global_materials[glass_code]['price']
                    else:
                        if '8' in glass_type: glass_price = 240000.0
                        elif '10' in glass_type: glass_price = 328000.0
                        elif '12' in glass_type: glass_price = 592900.0
                        else: glass_price = 200000.0
                    
            glass_area_ratio = 0.87
            glass_area = area * glass_area_ratio
            glass_cost = glass_area * glass_price
            
            sum_materials = profiles_cost + glass_cost + acc_cost
        pct_lab = door['percent_labor'] / 100.0
        
        if door.get('override_labor_cost') is not None:
            labor_cost_per_unit = door['override_labor_cost']
        else:
            fab_sel = indirect_selections.get('fabrication')
            if fab_sel:
                if fab_sel['custom_value'] is not None: labor_cost_per_unit = fab_sel['custom_value'] * area
                elif fab_sel['value_type'] == 'fixed': labor_cost_per_unit = fab_sel['value'] * area
                elif fab_sel['value_type'] == 'percent': labor_cost_per_unit = (fab_sel['value'] / 100.0) * sum_materials
                else: labor_cost_per_unit = pct_lab * (sum_materials / (1.0 - pct_lab) if pct_lab < 1.0 else sum_materials)
            else: labor_cost_per_unit = pct_lab * (sum_materials / (1.0 - pct_lab) if pct_lab < 1.0 else sum_materials)

        if door.get('override_installation_cost') is not None:
            installation_cost_per_unit = door['override_installation_cost']
        else:
            inst_sel = indirect_selections.get('installation')
            if inst_sel:
                if inst_sel['custom_value'] is not None: installation_cost_per_unit = (inst_sel['custom_value'] / 100.0) * sum_materials
                elif inst_sel['value_type'] == 'percent': installation_cost_per_unit = (inst_sel['value'] / 100.0) * sum_materials
                elif inst_sel['value_type'] == 'fixed': installation_cost_per_unit = inst_sel['value'] * area
                else: installation_cost_per_unit = 0.05 * sum_materials
            else: installation_cost_per_unit = 0.05 * sum_materials

        if door.get('override_transport_cost') is not None:
            transport_cost_per_unit = door['override_transport_cost']
        else:
            if trans_sel:
                if trans_sel['custom_value'] is not None:
                    val = trans_sel['custom_value']
                    transport_cost_per_unit = (val * area) / project_total_area if project_total_area > 0 else 0.0
                elif trans_sel['value_type'] == 'fixed':
                    val = trans_sel['value']
                    transport_cost_per_unit = (val * area) / project_total_area if project_total_area > 0 else 0.0
                elif trans_sel['value_type'] == 'percent': transport_cost_per_unit = (trans_sel['value'] / 100.0) * sum_materials
                else: transport_cost_per_unit = 0.0
            else: transport_cost_per_unit = 0.0

        cont_sel = indirect_selections.get('contingency')
        if cont_sel:
            if cont_sel['custom_value'] is not None: contingency_cost_per_unit = (cont_sel['custom_value'] / 100.0) * sum_materials
            elif cont_sel['value_type'] == 'percent': contingency_cost_per_unit = (cont_sel['value'] / 100.0) * sum_materials
            elif cont_sel['value_type'] == 'fixed': contingency_cost_per_unit = cont_sel['value']
            else: contingency_cost_per_unit = 0.02 * sum_materials
        else: contingency_cost_per_unit = 0.02 * sum_materials
            
        cost_per_unit = sum_materials + labor_cost_per_unit + installation_cost_per_unit + transport_cost_per_unit + contingency_cost_per_unit
        if door.get('price_per_m2') is not None and door['price_per_m2'] > 0:
            price_per_m2_rounded = door['price_per_m2']
            total_unit_cost_final = price_per_m2_rounded * area
        else:
            selling_price_raw = cost_per_unit * (1.0 + target_profit_margin / 100.0)
            price_per_m2 = selling_price_raw / area if area > 0 else 0.0
            price_per_m2_rounded = round(price_per_m2, -3)
            total_unit_cost_final = price_per_m2_rounded * area
            
        final_total_price = total_unit_cost_final * qty_sets
        results.append({
            'door_id': door_id, 'template_code': door['template_code'], 'code': door_code, 'name': door_name,
            'type': door_type, 'width': w, 'height': h, 'qty': qty_sets, 'area': area, 'total_area': total_area,
            'price_per_m2': price_per_m2_rounded, 'unit_price': total_unit_cost_final, 'total_price': final_total_price,
            'glass_type': glass_type, 'description': door.get('description') or '',
            'override_transport_cost': door.get('override_transport_cost'),
            'override_installation_cost': door.get('override_installation_cost'),
            'override_labor_cost': door.get('override_labor_cost'),
            'cost_per_unit': cost_per_unit, 'total_cost': cost_per_unit * qty_sets,
            'components': {
                'aluminum': profiles_cost, 'glass': glass_cost, 'auxiliary': acc_cost, 'materials_total': sum_materials,
                'labor': labor_cost_per_unit, 'installation': installation_cost_per_unit, 'transport': transport_cost_per_unit, 'contingency': contingency_cost_per_unit
            },
            'profiles': profiles_details, 'accessories': acc_details, 'glass_price_used': glass_price, 'glass_area_used': glass_area
        })
    conn.close()
    return results

def get_row_styles(ws, row_idx):
    styles = []
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row_idx, column=col)
        styles.append({
            'font': copy(cell.font),
            'fill': copy(cell.fill),
            'border': copy(cell.border),
            'alignment': copy(cell.alignment),
            'number_format': cell.number_format
        })
    row_height = ws.row_dimensions[row_idx].height
    return {'cells': styles, 'height': row_height}

def apply_row_styles(ws, row_idx, style_dict):
    styles = style_dict['cells']
    for col, style in enumerate(styles, 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.font = copy(style['font'])
        cell.fill = copy(style['fill'])
        cell.border = copy(style['border'])
        cell.alignment = copy(style['alignment'])
        cell.number_format = style['number_format']
    if style_dict['height'] is not None:
        ws.row_dimensions[row_idx].height = style_dict['height']

def get_door_description(item):
    name = item.get('name') or item.get('template_name') or ''
    system = item.get('system_name') or 'Nova System'
    glass = item.get('glass_type') or 'k8cl'
    brand = item.get('accessory_brand') or 'Draho'
    
    # Format glass type to readable Vietnamese
    glass_str = f"Kính {glass}"
    if 'cl' in glass.lower() or 'cường lực' in glass.lower():
        thickness_match = re.search(r'\d+', glass)
        thickness = thickness_match.group(0) if thickness_match else '8'
        glass_str = f"Kính trắng cường lực dày {thickness}mm"
    elif 'dán' in glass.lower() or 'an toàn' in glass.lower():
        thickness_match = re.search(r'\d+\.?\d*', glass)
        thickness = thickness_match.group(0) if thickness_match else '8.38'
        glass_str = f"Kính dán an toàn dày {thickness}mm"
    
    # Format accessory brand
    brand_str = f"Phụ kiện đồng bộ {brand}"
    if 'draho' in brand.lower():
        brand_str = "Phụ kiện bánh xe đôi, khóa âm đồng bộ Draho" if "lùa" in name.lower() else "Phụ kiện tay nắm, bản lề chữ A đồng bộ Draho"
        if "đi" in name.lower() or "cd" in name.lower() or "cửa đi" in name.lower():
            brand_str = "Phụ kiện khóa đa điểm, bản lề 3D đồng bộ Draho"
            
    desc = f"{name}\n- Nhóm hệ {system}\n- {glass_str}\n- {brand_str}"
    return desc

def generate_excel_report(project_id, template_path, output_path, split_output=False):
    """
    Generate Excel Report based on project calculations, copying template styles.

    When split_output is True, output_path is ignored and the function
    instead returns a (cost_wb, quote_wb) tuple of openpyxl Workbook objects
    ("Tổng hợp chi phí" and "Báo giá" as two independent workbooks), or
    (None, None) if the report could not be split.
    """
    print(f"Generating Excel Report from {template_path} to {output_path}")
    
    # Calculate project estimates
    calc_results = calculate_project_estimates(project_id)
    if not calc_results:
        print("No doors to estimate.")
        return (None, None) if split_output else False

    # Check if project has Opera BOM
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT has_opera_bom FROM projects WHERE id = %s", (project_id,))
    proj_row = cursor.fetchone()
    has_opera_bom = proj_row['has_opera_bom'] if proj_row and 'has_opera_bom' in proj_row else False
    
    opera_aluminum = []
    opera_glass = []
    opera_accessory = []
    opera_gasket = []
    opera_auxiliary = []
    
    if has_opera_bom:
        cursor.execute("""
        SELECT 
            pom.code,
            pom.name,
            pom.description,
            pom.quantity_unit,
            SUM(pom.quantity) as total_quantity,
            AVG(pom.unit_weight) as unit_weight,
            pom.color,
            pom.unit_price,
            m.category as catalog_category
        FROM project_opera_materials pom
        LEFT JOIN materials m ON pom.mapped_material_id = m.id
        WHERE pom.project_id = %s
        GROUP BY pom.code, pom.name, pom.description, pom.quantity_unit, pom.color, pom.unit_price, m.category
        ORDER BY pom.quantity_unit, pom.code
        """, (project_id,))
        opera_materials = [dict(row) for row in cursor.fetchall()]
        
        for mat in opera_materials:
            cat = mat['catalog_category']
            if not cat:
                code_lower = mat['code'].lower()
                name_lower = mat['name'].lower()
                unit_lower = mat['quantity_unit'].lower()
                if 'glass' in code_lower or 'kinh' in code_lower or 'kinh' in name_lower or 'gl' in code_lower:
                    cat = 'glass'
                elif 'gioang' in code_lower or 'gasket' in code_lower or 'gioăng' in name_lower:
                    cat = 'gasket'
                elif 'ac' in code_lower or 'pk' in code_lower or 'phu kien' in name_lower or 'accessory' in code_lower:
                    cat = 'accessory'
                elif 'al' in code_lower or 'nhom' in name_lower or mat['unit_weight'] is not None or unit_lower in ['m', 'meter', 'kg']:
                    cat = 'aluminum'
                else:
                    cat = 'other'
            
            if cat == 'aluminum':
                opera_aluminum.append(mat)
            elif cat == 'glass':
                opera_glass.append(mat)
            elif cat == 'accessory':
                opera_accessory.append(mat)
            elif cat == 'gasket':
                opera_gasket.append(mat)
            else:
                opera_auxiliary.append(mat)
    conn.close()
        
    # Open template workbook
    wb = openpyxl.load_workbook(template_path)
    
    # Update sheet DETAIL
    total_row = 52 # default fallback
    if 'DETAIL' in wb.sheetnames:
        ws_detail = wb['DETAIL']
        
        # 1. Read and save the styles of the template rows so we can apply them
        # Header Row style: from Row 9 of the template
        header_styles = get_row_styles(ws_detail, 9)
        # Data Row style: from Row 10 of the template
        data_styles = get_row_styles(ws_detail, 10)
        # Total Row style: from Row 52 of the template
        total_styles = get_row_styles(ws_detail, 52)
        # Ghi chu Row style: from Row 54 and 55 of the template
        ghichu_header_styles = get_row_styles(ws_detail, 54)
        ghichu_body_styles = get_row_styles(ws_detail, 55)
        
        # Load the project name to put in the header
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT name FROM projects WHERE id = %s", (project_id,))
        proj_row = cursor.fetchone()
        project_name = proj_row[0] if proj_row else "GOLDEN CITY"
        conn.close()
        
        # Set project name in Row 2
        ws_detail.cell(row=2, column=1, value=f"Dự án: {project_name.upper()}")
        
        # 1.5. Unmerge all merged ranges from row 9 downwards to avoid write restrictions
        merged_ranges = list(ws_detail.merged_cells.ranges)
        for r_range in merged_ranges:
            if r_range.min_row >= 9:
                ws_detail.unmerge_cells(str(r_range))
        
        # 2. Clear all rows from row 9 to the max row of the worksheet and clear old groupings
        # (This avoids any leftovers and incorrect outlines from the template)
        original_max_row = ws_detail.max_row
        for r in range(9, original_max_row + 1):
            ws_detail.row_dimensions[r].hidden = False # Unhide
            ws_detail.row_dimensions[r].outline_level = 0 # Clear group Outline level
            for col in range(1, ws_detail.max_column + 1):
                cell = ws_detail.cell(row=r, column=col)
                if cell.__class__.__name__ == 'Cell':
                    cell.value = None
                
        # 3. Group calc_results by door type
        grouped_items = {}
        for item in calc_results:
            t = item.get('type') or 'CỬA KHÁC'
            if t not in grouped_items:
                grouped_items[t] = []
            grouped_items[t].append(item)
            
        def type_sort_key(t):
            t_lower = t.lower()
            if 'sổ' in t_lower:
                return (0, t)
            elif 'đi' in t_lower:
                return (1, t)
            elif 'vách' in t_lower:
                return (2, t)
            else:
                return (3, t)
                
        sorted_types = sorted(grouped_items.keys(), key=type_sort_key)
        
        # Write categories and items
        current_row = 9
        category_letters = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']
        
        for cat_idx, cat_type in enumerate(sorted_types):
            cat_letter = category_letters[cat_idx] if cat_idx < len(category_letters) else str(cat_idx + 1)
            cat_items = grouped_items[cat_type]
            
            # Write Category Header Row
            ws_detail.cell(row=current_row, column=1, value=cat_letter)
            ws_detail.cell(row=current_row, column=2, value=cat_type.upper())
            apply_row_styles(ws_detail, current_row, header_styles)
            
            # Merge columns 2 to 7 for Category Header
            try:
                ws_detail.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=7)
            except Exception as e:
                print(f"Warning merging cells at row {current_row}: {e}")
            
            start_r = current_row + 1
            end_r = current_row + len(cat_items)
            
            # Update Category Header subtotal formulas
            # Column H: =SUBTOTAL(9,H{start}:H{end})
            # Column J: =SUBTOTAL(9,J{start}:J{end})
            ws_detail.cell(row=current_row, column=8, value=f"=SUBTOTAL(9,H{start_r}:H{end_r})")
            ws_detail.cell(row=current_row, column=10, value=f"=SUBTOTAL(9,J{start_r}:J{end_r})")
            
            current_row += 1
            
            # Write Category Data Rows
            for item_idx, item in enumerate(cat_items, 1):
                # Column A: STT (integer)
                ws_detail.cell(row=current_row, column=1, value=item_idx)
                
                # Column B: Description
                desc = item.get('description')
                if not desc or not desc.strip():
                    desc = get_door_description(item)
                ws_detail.cell(row=current_row, column=2, value=desc)
                # Enable wrap text for description
                ws_detail.cell(row=current_row, column=2).alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='center')
                
                # Column C: Ký hiệu
                ws_detail.cell(row=current_row, column=3, value=item.get('code') or item.get('template_code'))
                
                # Column D: Rộng (width in meters)
                w_val = item.get('width')
                w_m = w_val if w_val < 10.0 else w_val / 1000.0
                ws_detail.cell(row=current_row, column=4, value=w_m)
                
                # Column E: Cao (height in meters)
                h_val = item.get('height')
                h_m = h_val if h_val < 10.0 else h_val / 1000.0
                ws_detail.cell(row=current_row, column=5, value=h_m)
                
                # Column F: Đơn vị
                ws_detail.cell(row=current_row, column=6, value='m2')
                
                # Column G: Số lượng
                ws_detail.cell(row=current_row, column=7, value=item.get('qty'))
                
                # Column H: Tổng cộng (m2) = D*E*G
                ws_detail.cell(row=current_row, column=8, value=f"=D{current_row}*E{current_row}*G{current_row}")
                
                # Column I: Đơn giá/m2
                ws_detail.cell(row=current_row, column=9, value=item.get('price_per_m2'))
                
                # Column J: Thành tiền = H*I
                ws_detail.cell(row=current_row, column=10, value=f"=H{current_row}*I{current_row}")
                
                apply_row_styles(ws_detail, current_row, data_styles)
                current_row += 1
                
            # Group the data rows under this category dynamically
            if end_r >= start_r:
                ws_detail.row_dimensions.group(start_r, end_r, outline_level=1, hidden=False)
                
        # 4. Write Total Row
        total_row = current_row
        ws_detail.cell(row=total_row, column=1, value="Tổng cộng:")
        
        # Formulas for Total Row:
        # H: `=SUBTOTAL(9,H9:H{total_row-1})`
        # I: `=IFERROR(J{total_row}/H{total_row},"")`
        # J: `=SUBTOTAL(9,J9:J{total_row-1})`
        ws_detail.cell(row=total_row, column=8, value=f"=SUBTOTAL(9,H9:H{total_row-1})")
        ws_detail.cell(row=total_row, column=9, value=f'=IFERROR(J{total_row}/H{total_row},"")')
        ws_detail.cell(row=total_row, column=10, value=f"=SUBTOTAL(9,J9:J{total_row-1})")
        
        apply_row_styles(ws_detail, total_row, total_styles)
        current_row += 1
        
        # 5. Write Ghi chú section
        current_row += 1 # Empty row
        
        ws_detail.cell(row=current_row, column=1, value="Ghi chú:")
        apply_row_styles(ws_detail, current_row, ghichu_header_styles)
        current_row += 1
        
        # Body of Ghi chú - write as plain strings to prevent sheet reference errors
        ghi_chu_lines = [
            "- Sản phẩm sử dụng hệ nhôm hoàn thiện sơn tĩnh điện theo tiêu chuẩn AAMA2604, bảo hành sơn 20 năm, màu sơn solid (không bao gồm màu sơn metalic).",
            "- Đơn giá áp dụng cho 1 màu sơn cho toàn dự án, trường hợp dự án có nhiều màu sẽ được tính toán phụ phí và thống nhất giữa các bên.",
            "- Kích thước nghiệm thu là kích thước ô chờ.",
            "- Đối với các hạng mục có hình dạng dị hình (vòm cong, tam giác, bình hành…), khối lượng được quy về hình chữ nhật bao quanh.",
            "- Khối lượng là tạm tính.",
            "- Báo giá không bao gồm thí nghiệm mô hình."
        ]
        
        for line in ghi_chu_lines:
            ws_detail.cell(row=current_row, column=1, value=line)
            apply_row_styles(ws_detail, current_row, ghichu_body_styles)
            current_row += 1

    # Update sheet CPHoanThien
    if 'CPHoanThien' in wb.sheetnames:
        ws_summary = wb['CPHoanThien']
        
        # 1. Fetch cost overhead percentages from database
        conn_cost = get_db_connection()
        cursor_cost = conn_cost.cursor()
        
        p_company = 2.0
        p_contingency = 2.0
        p_warranty = 1.5
        p_other = 1.0
        
        try:
            cursor_cost.execute("SELECT pct_company, pct_contingency, pct_warranty, pct_other FROM projects WHERE id = %s", (project_id,))
            proj_cost = cursor_cost.fetchone()
            if proj_cost:
                p_company = proj_cost['pct_company'] if proj_cost['pct_company'] is not None else 2.0
                p_contingency = proj_cost['pct_contingency'] if proj_cost['pct_contingency'] is not None else 2.0
                p_warranty = proj_cost['pct_warranty'] if proj_cost['pct_warranty'] is not None else 1.5
                p_other = proj_cost['pct_other'] if proj_cost['pct_other'] is not None else 1.0
        except Exception as e:
            print(f"Warning: Could not fetch project cost percentages: {e}")
        finally:
            cursor_cost.close()
            conn_cost.close()

        # 2. Compute project totals and scale factor based on door areas
        total_project_area = sum(item['total_area'] for item in calc_results)
        # Template sample total door area is 12156.0 m2 (from row 14 col E of original template)
        scale_factor = total_project_area / 12156.0 if total_project_area > 0 else 0.0

        # 3. Categorize door areas for labor cost breakdown
        area_di_lua = 0.0
        area_di_quay = 0.0
        area_so_lua = 0.0
        area_vach_kinh = 0.0
        
        for item in calc_results:
            t_code = item['template_code'].upper()
            t_name = item['name'].upper()
            t_area = item['total_area']
            
            if 'VKT' in t_code or 'VÁCH' in t_name:
                area_vach_kinh += t_area
            elif 'CỬA ĐI' in item['type'].upper() or 'CD' in t_code:
                if 'LÙA' in t_name or 'CL' in t_code or 'SL' in t_code:
                    area_di_lua += t_area
                else:
                    area_di_quay += t_area
            elif 'CỬA SỔ' in item['type'].upper() or 'CS' in t_code:
                if 'LÙA' in t_name or 'SL' in t_code:
                    area_so_lua += t_area
                else:
                    area_so_lua += t_area # Fallback window area to window sliding rate

        # 4. Aggregate actual project materials
        project_aluminum = {} # code -> total_weight (kg)
        project_accessories = {} # code -> total_qty
        project_glass = {} # code -> total_area (m2)

        for item in calc_results:
            qty_sets = item['qty']
            
            # Aluminum
            for p in item.get('profiles', []):
                p_code = p['code'].strip().upper()
                p_weight = p['total_weight'] * qty_sets
                project_aluminum[p_code] = project_aluminum.get(p_code, 0.0) + p_weight
                
            # Accessories
            for a in item.get('accessories', []):
                a_code = a['code'].strip().upper()
                a_qty = a['qty'] * qty_sets
                project_accessories[a_code] = project_accessories.get(a_code, 0.0) + a_qty
                
            # Glass
            g_type = item['glass_type'].strip().lower()
            g_area = item.get('glass_area_used', item['area'] * 0.87) * qty_sets
            matched_glass = 'bk'
            if '8' in g_type:
                matched_glass = 'k8cl'
            elif '10' in g_type:
                matched_glass = 'k10cl'
            project_glass[matched_glass] = project_glass.get(matched_glass, 0.0) + g_area

        # 5. Load material prices from database for precise cost calculation
        conn_mats = get_db_connection()
        cursor_mats = conn_mats.cursor()
        project_prices = {}
        book_prices = {}
        global_prices = {}
        
        try:
            # Get project price book info
            cursor_mats.execute("SELECT price_book_id FROM projects WHERE id = %s", (project_id,))
            proj_row = cursor_mats.fetchone()
            price_book_id = proj_row[0] if proj_row else None

            # Project-specific prices
            cursor_mats.execute("SELECT material_code, price FROM project_material_prices WHERE project_id = %s", (project_id,))
            project_prices = {row[0].strip().upper(): float(row[1]) for row in cursor_mats.fetchall() if row[0] and row[1] is not None}
            
            # Price book prices
            if price_book_id:
                cursor_mats.execute("SELECT material_code, price FROM material_price_book_items WHERE price_book_id = %s", (price_book_id,))
                book_prices = {row[0].strip().upper(): float(row[1]) for row in cursor_mats.fetchall() if row[0] and row[1] is not None}

            # Global default prices
            cursor_mats.execute("SELECT code, default_price FROM materials")
            global_prices = {row[0].strip().upper(): float(row[1]) for row in cursor_mats.fetchall() if row[0] and row[1] is not None}
        except Exception as e:
            print(f"Warning: Could not fetch database material prices: {e}")
        finally:
            cursor_mats.close()
            conn_mats.close()

        # 5.5. Read default values first so we can compute rates and scale others
        default_values = {}
        for r in range(1, 150):
            stt = ws_summary.cell(row=r, column=1).value
            stt_str = str(stt).strip() if stt is not None else ""
            val = ws_summary.cell(row=r, column=2).value
            val_str = str(val).strip() if val is not None else ""
            
            qty_val = ws_summary.cell(row=r, column=5).value
            price_val = ws_summary.cell(row=r, column=6).value
            total_val = ws_summary.cell(row=r, column=7).value
            
            default_values[r] = {
                'stt': stt_str,
                'val': val_str,
                'qty': float(qty_val) if isinstance(qty_val, (int, float)) else 0.0,
                'price': float(price_val) if isinstance(price_val, (int, float)) else 0.0,
                'total': float(total_val) if isinstance(total_val, (int, float)) else 0.0
            }

        # Helper function to find the actual price of a material
        def get_actual_price(mat_code, d_row):
            code_upper = str(mat_code).strip().upper()
            if code_upper in project_prices:
                return project_prices[code_upper]
            if code_upper in book_prices:
                return book_prices[code_upper]
            if code_upper in global_prices:
                return global_prices[code_upper]
            # Fallback to template-defined price (total / qty)
            if d_row['qty'] > 0:
                return d_row['total'] / d_row['qty']
            return d_row['price']

        # 6. Update CPHoanThien sheet rows dynamically
        for r in range(1, 150):
            d = default_values.get(r)
            if not d or not d['val']:
                continue
                
            val_str = d['val'].strip().lower()
            stt_str = d['stt']
            
            # Check for DOANH THU to map total revenue (only the main revenue row with STT 'A')
            if 'doanh thu' in val_str and stt_str == 'A':
                ws_summary.cell(row=r, column=7, value=f"=DETAIL!J{total_row}")
                
            # Fill cost overhead percentages in Column F (Column 6)
            elif 'chi phí công ty' in val_str or 'chi phi cong ty' in val_str:
                ws_summary.cell(row=r, column=6, value=p_company / 100.0)
            elif 'dự phòng phí' in val_str or 'du phong phi' in val_str:
                ws_summary.cell(row=r, column=6, value=p_contingency / 100.0)
            elif 'dự phòng bảo hành' in val_str or 'du phong bao hanh' in val_str:
                ws_summary.cell(row=r, column=6, value=p_warranty / 100.0)
            elif 'chi phí khác' in val_str or 'chi phi khac' in val_str:
                ws_summary.cell(row=r, column=6, value=p_other / 100.0)
                
            # Direct labor cost breakdown rows
            elif 'nhân công gia công' in val_str or 'nhân công lắp đặt' in val_str or 'chi phí vệ sinh' in val_str:
                # Find labor category rate
                rate = d['total'] / d['qty'] if d['qty'] > 0 else d['price']
                
                # Determine door area type
                door_area = 0.0
                if 'cửa đi lùa' in val_str:
                    door_area = area_di_lua
                elif 'cửa đi mở quay' in val_str:
                    door_area = area_di_quay
                elif 'cửa sổ lùa' in val_str:
                    door_area = area_so_lua
                elif 'vách kính' in val_str:
                    door_area = area_vach_kinh
                    
                ws_summary.cell(row=r, column=5, value=door_area)
                ws_summary.cell(row=r, column=6, value=rate)
                ws_summary.cell(row=r, column=7, value=f"=E{r}*F{r}")
                
            # Aluminum cost rows
            elif r >= 25 and r <= 57: # Aluminum detailed items
                if has_opera_bom:
                    idx = r - 25
                    if idx < len(opera_aluminum):
                        mat = opera_aluminum[idx]
                        ws_summary.cell(row=r, column=2, value=mat['code'])
                        ws_summary.cell(row=r, column=3, value=mat['name'])
                        unit = 'Kg' if mat['unit_weight'] and mat['quantity_unit'] in ['m', 'meter'] else mat['quantity_unit']
                        qty = mat['total_quantity'] * mat['unit_weight'] if (mat['unit_weight'] and mat['quantity_unit'] in ['m', 'meter']) else mat['total_quantity']
                        ws_summary.cell(row=r, column=4, value=unit)
                        ws_summary.cell(row=r, column=5, value=qty)
                        ws_summary.cell(row=r, column=6, value=mat['unit_price'] or 0.0)
                        ws_summary.cell(row=r, column=7, value=f"=E{r}*F{r}")
                    else:
                        for col in [2, 3, 4, 5, 6, 7]:
                            ws_summary.cell(row=r, column=col, value=None)
                else:
                    mat_code = d['val'].strip().upper()
                    act_weight = project_aluminum.get(mat_code, 0.0)
                    price = get_actual_price(mat_code, d)
                    ws_summary.cell(row=r, column=5, value=act_weight)
                    ws_summary.cell(row=r, column=6, value=price)
                    ws_summary.cell(row=r, column=7, value=f"=E{r}*F{r}")
                
            # Glass cost rows
            elif r >= 59 and r <= 64: # Glass detailed items
                if has_opera_bom:
                    idx = r - 59
                    if idx < len(opera_glass):
                        mat = opera_glass[idx]
                        ws_summary.cell(row=r, column=2, value=mat['code'])
                        ws_summary.cell(row=r, column=3, value=mat['name'])
                        ws_summary.cell(row=r, column=4, value=mat['quantity_unit'])
                        ws_summary.cell(row=r, column=5, value=mat['total_quantity'])
                        ws_summary.cell(row=r, column=6, value=mat['unit_price'] or 0.0)
                        ws_summary.cell(row=r, column=7, value=f"=E{r}*F{r}")
                    else:
                        for col in [2, 3, 4, 5, 6, 7]:
                            ws_summary.cell(row=r, column=col, value=None)
                else:
                    mat_code = d['val'].strip().lower()
                    if mat_code in ['k8cl', 'k10cl', 'bk']:
                        act_area = project_glass.get(mat_code, 0.0)
                        price = get_actual_price(mat_code, d)
                        ws_summary.cell(row=r, column=5, value=act_area)
                        ws_summary.cell(row=r, column=6, value=price)
                        ws_summary.cell(row=r, column=7, value=f"=E{r}*F{r}")
                    else: # Glass processing services (scale by scale_factor)
                        act_qty = d['qty'] * scale_factor
                        price = get_actual_price(mat_code, d)
                        ws_summary.cell(row=r, column=5, value=act_qty)
                        ws_summary.cell(row=r, column=6, value=price)
                        ws_summary.cell(row=r, column=7, value=f"=E{r}*F{r}")
                    
            # Accessory cost rows
            elif r >= 66 and r <= 80: # Accessories detailed items
                if has_opera_bom:
                    idx = r - 66
                    if idx < len(opera_accessory):
                        mat = opera_accessory[idx]
                        ws_summary.cell(row=r, column=2, value=mat['code'])
                        ws_summary.cell(row=r, column=3, value=mat['name'])
                        ws_summary.cell(row=r, column=4, value=mat['quantity_unit'])
                        ws_summary.cell(row=r, column=5, value=mat['total_quantity'])
                        ws_summary.cell(row=r, column=6, value=mat['unit_price'] or 0.0)
                        ws_summary.cell(row=r, column=7, value=f"=E{r}*F{r}")
                    else:
                        for col in [2, 3, 4, 5, 6, 7]:
                            ws_summary.cell(row=r, column=col, value=None)
                else:
                    mat_code = d['val'].strip().upper()
                    act_qty = project_accessories.get(mat_code, 0.0)
                    price = get_actual_price(mat_code, d)
                    ws_summary.cell(row=r, column=5, value=act_qty)
                    ws_summary.cell(row=r, column=6, value=price)
                    ws_summary.cell(row=r, column=7, value=f"=E{r}*F{r}")
                
            # Gasket, auxiliary, transport (scale qty & total)
            elif (r >= 82 and r <= 83) or (r >= 85 and r <= 88) or r == 90:
                if has_opera_bom and r != 90:
                    if r >= 82 and r <= 83:
                        idx = r - 82
                        mats_list = opera_gasket
                    else:
                        idx = r - 85
                        mats_list = opera_auxiliary
                        
                    if idx < len(mats_list):
                        mat = mats_list[idx]
                        ws_summary.cell(row=r, column=2, value=mat['code'])
                        ws_summary.cell(row=r, column=3, value=mat['name'])
                        ws_summary.cell(row=r, column=4, value=mat['quantity_unit'])
                        ws_summary.cell(row=r, column=5, value=mat['total_quantity'])
                        ws_summary.cell(row=r, column=6, value=mat['unit_price'] or 0.0)
                        ws_summary.cell(row=r, column=7, value=f"=E{r}*F{r}")
                    else:
                        for col in [2, 3, 4, 5, 6, 7]:
                            ws_summary.cell(row=r, column=col, value=None)
                else:
                    act_qty = d['qty'] * scale_factor
                    price = get_actual_price(d['val'], d)
                    ws_summary.cell(row=r, column=5, value=act_qty)
                    ws_summary.cell(row=r, column=6, value=price)
                    ws_summary.cell(row=r, column=7, value=f"=E{r}*F{r}")
                
            # Equipment cost (row 99) and management overhead (row 118) (scale qty & total)
            elif r == 99 or r == 118:
                act_qty = d['qty'] * scale_factor
                price = d['total'] / d['qty'] if d['qty'] > 0 else d['price']
                ws_summary.cell(row=r, column=5, value=act_qty)
                ws_summary.cell(row=r, column=6, value=price)
                ws_summary.cell(row=r, column=7, value=f"=E{r}*F{r}")

    # Update detail breakdown sheets (CSL-50.01, CSL-50.02, CSL-50.03...)
    for item in calc_results:
        sheet_name = item['template_code']
        if sheet_name in wb.sheetnames:
            ws_sheet = wb[sheet_name]
            # Update header info
            ws_sheet.cell(row=4, column=6, value=item['price_per_m2'])
            ws_sheet.cell(row=5, column=6, value=item['area'])
            ws_sheet.cell(row=6, column=6, value=item['unit_price'])
            ws_sheet.cell(row=7, column=6, value=item['qty'])
            
            # Update profile table values
            profiles = item['profiles']
            for p in profiles:
                for r in range(12, 40):
                    cell_code = ws_sheet.cell(row=r, column=2).value
                    if cell_code == p['code']:
                        ws_sheet.cell(row=r, column=4, value=p['length'])
                        ws_sheet.cell(row=r, column=7, value=p['total_weight'])
                        ws_sheet.cell(row=r, column=9, value=p['total_price'])
                        ws_sheet.cell(row=r, column=8, value=p['unit_price'])
                        break
                        
    # Remove unused sheets
    active_templates = {item['template_code'] for item in calc_results}
    sheets_to_remove = []
    for sheet_name in wb.sheetnames:
        is_door_sheet = any(sheet_name.startswith(prefix) for prefix in ['CSL-', 'CDL-', 'CSB-', 'CDMQ-', 'VKT'])
        if is_door_sheet and sheet_name not in active_templates:
            sheets_to_remove.append(sheet_name)

    for sheet_name in sheets_to_remove:
        try:
            wb.remove(wb[sheet_name])
            print(f"Removed unused sheet: {sheet_name}")
        except Exception as e:
            print(f"Error removing sheet {sheet_name}: {e}")

    if split_output:
        return _split_report_workbook(wb, total_row, calc_results)

    # Save output
    wb.save(output_path)
    print(f"Excel report saved successfully to {output_path}")
    return True


DOOR_SHEET_PREFIXES = ('CSL-', 'CDL-', 'CSB-', 'CDMQ-', 'VKT')


def _split_report_workbook(wb, total_row, calc_results):
    """
    Split the already-built combined report workbook into two standalone
    workbooks: a cost-summary workbook ("Tổng hợp chi phí") and a
    customer-facing quote workbook ("Báo giá"), per the requirement that
    these be two separate deliverable files.

    Returns (cost_wb, quote_wb) as openpyxl Workbook objects, or (None, None)
    if the required sheets are missing.
    """
    if 'DETAIL' not in wb.sheetnames:
        return None, None

    revenue_total = sum(float(item.get('total_price') or 0) for item in calc_results)

    # Save the fully-built workbook to a temp buffer and reload it twice so
    # each split copy has its own independent set of sheets/styles.
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    cost_wb = openpyxl.load_workbook(buf)
    buf.seek(0)
    quote_wb = openpyxl.load_workbook(buf)

    # --- Cost summary workbook: keep CPHoanThien + door breakdown sheets,
    # drop DETAIL and freeze the cross-sheet revenue formula to a static
    # value so the workbook stays self-contained without DETAIL.
    if 'CPHoanThien' in cost_wb.sheetnames:
        ws_summary = cost_wb['CPHoanThien']
        target_formula = f"=DETAIL!J{total_row}"
        for row in ws_summary.iter_rows():
            for cell in row:
                if cell.value == target_formula:
                    cell.value = revenue_total
    if 'DETAIL' in cost_wb.sheetnames:
        cost_wb.remove(cost_wb['DETAIL'])
    if cost_wb.sheetnames:
        cost_wb.active = 0

    # --- Quote workbook: keep DETAIL only, drop CPHoanThien and per-door
    # breakdown sheets so the customer receives just the quote table.
    for sheet_name in list(quote_wb.sheetnames):
        if sheet_name == 'CPHoanThien' or sheet_name.startswith(DOOR_SHEET_PREFIXES):
            quote_wb.remove(quote_wb[sheet_name])
    if 'DETAIL' in quote_wb.sheetnames:
        quote_wb.active = quote_wb.sheetnames.index('DETAIL')

    return cost_wb, quote_wb

def consolidate_aluminum_orders(file_info_list: list, output_path: str) -> bool:
    """
    Consolidate aluminum profile orders from multiple Opera Excel files.
    file_info_list: List of dicts, each containing:
        - 'path': absolute path to the uploaded file
        - 'original_name': original filename (e.g., 'Project_A_Opt.xls')
    output_path: Path to save the consolidated Excel report.
    """
    import pandas as pd
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    print(f"Consolidating {len(file_info_list)} aluminum order files to {output_path}")

    # Column helper mapping
    def find_column(df, possible_names):
        for col in df.columns:
            if str(col).strip().lower() in [name.lower() for name in possible_names]:
                return col
        return None

    possible_codes = ['Code', 'Part Number', 'Mã nhôm', 'Mã vật tư', 'Mã', 'PartNo', 'Item Code']
    possible_lengths = ['Length', 'Chiều dài', 'L', 'Kích thước', 'Dài', 'Length(mm)', 'Cut Length']
    possible_pieces = ['Pieces', 'Quantity', 'Qty', 'Số lượng', 'Sl', 'Số thanh', 'Pcs', 'QuantityUnit']
    possible_colors = ['Color', 'Màu', 'Màu sắc', 'Sơn', 'Finish']
    possible_descriptions = ['Description', 'Name', 'Mô tả', 'Tên', 'Ghi chú', 'Part Description']
    possible_weights = ['Unit Weight', 'Weight', 'Trọng lượng', 'Đơn trọng', 'Tỷ trọng', 'Weight/m']
    possible_units = ['QuantityUnit', 'Unit', 'Đơn vị', 'ĐVT']

    consolidated_data = {} # Key: (code, length, color) -> Details
    detailed_sources = []  # List of dicts to store trace details

    for file_info in file_info_list:
        file_path = file_info['path']
        orig_name = file_info['original_name']
        project_label = os.path.splitext(orig_name)[0]

        try:
            # Check extension to load correctly
            file_ext = os.path.splitext(file_path)[1].lower()
            if file_ext == '.xls':
                df = pd.read_excel(file_path, engine='xlrd')
            else:
                df = pd.read_excel(file_path)
        except Exception as e:
            print(f"Error reading file {orig_name}: {e}")
            continue

        # Clean column names
        df.columns = [str(c).strip() for c in df.columns]

        # Find mapped columns
        code_col = find_column(df, possible_codes)
        length_col = find_column(df, possible_lengths)
        pieces_col = find_column(df, possible_pieces)
        color_col = find_column(df, possible_colors)
        desc_col = find_column(df, possible_descriptions)
        weight_col = find_column(df, possible_weights)
        unit_col = find_column(df, possible_units)

        # We need at least Code and Pieces columns to proceed
        if not code_col or not pieces_col:
            print(f"File {orig_name} is missing core columns (Code or Pieces). Mapped columns: Code={code_col}, Pieces={pieces_col}. Skipping.")
            continue

        for idx, row in df.iterrows():
            code = str(row.get(code_col, '')).strip()
            if not code or code.lower() == 'nan' or 'thay the' in code.lower():
                continue

            # Parse quantity (Pieces)
            try:
                pieces = float(row.get(pieces_col, 0))
            except:
                pieces = 0.0
            if pd.isna(pieces) or pieces <= 0:
                continue

            # Parse Length
            try:
                length = float(row.get(length_col, 0))
            except:
                length = 0.0
            if pd.isna(length):
                length = 0.0

            # Check if this row is a profile (aluminum bar)
            # Conditions: Unit is 'pc' or Length is a large number (typically mm: e.g., 5000, 5800, 6000)
            # Accessories are usually short or have no length column.
            # We filter for Length >= 1000 to capture actual profile bars (1m to 6m+)
            unit_val = str(row.get(unit_col, 'pc')).strip().lower() if unit_col else 'pc'
            
            is_profile = False
            if length >= 1000:
                is_profile = True
            elif unit_val == 'pc' and length >= 1000:
                is_profile = True

            if not is_profile:
                continue

            # Color
            color = str(row.get(color_col, '')).strip() if color_col else ''
            if not color or color.lower() == 'nan':
                color = "Tiêu chuẩn"

            # Description
            desc = str(row.get(desc_col, '')).strip() if desc_col else ''
            if not desc or desc.lower() == 'nan':
                desc = "Thanh nhôm profile"

            # Unit Weight
            try:
                unit_weight = float(row.get(weight_col, 0))
            except:
                unit_weight = 0.0
            if pd.isna(unit_weight):
                unit_weight = 0.0

            key = (code, length, color)

            # Store consolidated sum
            if key not in consolidated_data:
                consolidated_data[key] = {
                    'code': code,
                    'length': length,
                    'color': color,
                    'description': desc,
                    'pieces': 0.0,
                    'unit_weight': unit_weight,
                    'sources': {}
                }
            
            consolidated_data[key]['pieces'] += pieces
            # Keep the first valid description or unit weight
            if desc and consolidated_data[key]['description'] == "Thanh nhôm profile":
                consolidated_data[key]['description'] = desc
            if unit_weight > 0 and consolidated_data[key]['unit_weight'] == 0:
                consolidated_data[key]['unit_weight'] = unit_weight

            # Track source breakdown
            consolidated_data[key]['sources'][project_label] = consolidated_data[key]['sources'].get(project_label, 0.0) + pieces

            # Also save to detailed sources for the second sheet
            detailed_sources.append({
                'source_file': project_label,
                'code': code,
                'description': desc,
                'length': length,
                'color': color,
                'pieces': pieces,
                'unit_weight': unit_weight,
                'total_weight': (pieces * length / 1000.0 * unit_weight) if unit_weight > 0 else 0.0
            })

    if not consolidated_data:
        print("No valid aluminum profile data found to consolidate.")
        return False, []

    # Convert to list and sort
    sorted_items = sorted(consolidated_data.values(), key=lambda x: (x['code'], -x['length'], x['color']))

    # Create Workbook
    wb = openpyxl.Workbook()
    
    # -------------------------------------------------------------------------
    # Sheet 1: CONSOLIDATED ORDER (TỔNG HỢP ĐẶT HÀNG)
    # -------------------------------------------------------------------------
    ws_main = wb.active
    ws_main.title = "TONG HOP DAT HANG"
    ws_main.views.sheetView[0].showGridLines = True

    # Typography & Palette
    font_family = "Segoe UI"
    navy_dark = "1B365D" # Primary Navy
    green_accent = "2DB34B" # Accent Green
    gray_light = "F1F5F9" # Zebra Row
    border_color = "CBD5E1" # Slate Border

    title_font = Font(name=font_family, size=16, bold=True, color=navy_dark)
    subtitle_font = Font(name=font_family, size=10, italic=True, color="555555")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    data_font = Font(name=font_family, size=11, color="000000")
    total_font = Font(name=font_family, size=11, bold=True, color="000000")

    header_fill = PatternFill(start_color=navy_dark, end_color=navy_dark, fill_type="solid")
    zebra_fill = PatternFill(start_color=gray_light, end_color=gray_light, fill_type="solid")
    total_fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")

    thin_border = Border(
        left=Side(style='thin', color=border_color),
        right=Side(style='thin', color=border_color),
        top=Side(style='thin', color=border_color),
        bottom=Side(style='thin', color=border_color)
    )
    
    double_bottom_border = Border(
        left=Side(style='thin', color=border_color),
        right=Side(style='thin', color=border_color),
        top=Side(style='thin', color=border_color),
        bottom=Side(style='double', color="000000")
    )

    # Write Titles
    ws_main.cell(row=2, column=1, value="NOVALAND E&C").font = Font(name=font_family, size=12, bold=True, color=green_accent)
    ws_main.cell(row=3, column=1, value="BẢNG TỔNG HỢP ĐẶT HÀNG VẬT TƯ NHÔM (OPERA CONSOLIDATED)").font = title_font
    ws_main.cell(row=4, column=1, value=f"Gom từ {len(file_info_list)} file tối ưu hóa dự án | Ngày tổng hợp: {pd.Timestamp.now().strftime('%d/%m/%Y %H:%M')}").font = subtitle_font

    # Headers
    headers = [
        "STT", "Mã Nhôm (Code)", "Mô Tả Vật Tư (Description)", 
        "Chiều Dài (mm)", "Màu Sắc (Color)", "ĐVT", 
        "Số Lượng (Thanh)", "Đơn Trọng (kg/m)", "Khối Lượng (kg)", 
        "Nguồn Gốc Chi Tiết (Breakdown)"
    ]
    
    header_row = 6
    for col_idx, h in enumerate(headers, 1):
        cell = ws_main.cell(row=header_row, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    
    ws_main.row_dimensions[header_row].height = 28

    # Write Data
    current_row = 7
    for idx, item in enumerate(sorted_items, 1):
        ws_main.cell(row=current_row, column=1, value=idx).alignment = Alignment(horizontal="center")
        
        c_cell = ws_main.cell(row=current_row, column=2, value=item['code'])
        c_cell.font = Font(name=font_family, size=11, bold=True)
        c_cell.alignment = Alignment(horizontal="left")
        
        ws_main.cell(row=current_row, column=3, value=item['description']).alignment = Alignment(horizontal="left")
        
        l_cell = ws_main.cell(row=current_row, column=4, value=item['length'])
        l_cell.number_format = "#,##0"
        l_cell.alignment = Alignment(horizontal="right")
        
        ws_main.cell(row=current_row, column=5, value=item['color']).alignment = Alignment(horizontal="center")
        ws_main.cell(row=current_row, column=6, value="Thanh").alignment = Alignment(horizontal="center")
        
        q_cell = ws_main.cell(row=current_row, column=7, value=item['pieces'])
        q_cell.number_format = "#,##0"
        q_cell.alignment = Alignment(horizontal="right")
        
        w_cell = ws_main.cell(row=current_row, column=8, value=item['unit_weight'])
        w_cell.number_format = "0.000"
        w_cell.alignment = Alignment(horizontal="right")
        
        # Formula for Total Weight = Pieces * Length / 1000 * Unit_Weight
        tw_cell = ws_main.cell(row=current_row, column=9, value=f"=G{current_row}*D{current_row}/1000*H{current_row}")
        tw_cell.number_format = "#,##0.00"
        tw_cell.alignment = Alignment(horizontal="right")

        # Source breakdown string: "ProjA: 5, ProjB: 10"
        source_str = ", ".join([f"{k}: {int(v) if v.is_integer() else v} thanh" for k, v in item['sources'].items()])
        ws_main.cell(row=current_row, column=10, value=source_str).alignment = Alignment(horizontal="left", wrap_text=True)

        # Apply basic styles and borders
        for col_idx in range(1, len(headers) + 1):
            cell = ws_main.cell(row=current_row, column=col_idx)
            cell.font = data_font
            cell.border = thin_border
            # Zebra striping
            if idx % 2 == 0:
                cell.fill = zebra_fill
                
        ws_main.row_dimensions[current_row].height = 20
        current_row += 1

    # Write Total Row
    total_row_idx = current_row
    ws_main.cell(row=total_row_idx, column=1, value="Tổng cộng:").alignment = Alignment(horizontal="center")
    ws_main.cell(row=total_row_idx, column=7, value=f"=SUM(G7:G{total_row_idx-1})").number_format = "#,##0"
    ws_main.cell(row=total_row_idx, column=7).alignment = Alignment(horizontal="right")
    ws_main.cell(row=total_row_idx, column=9, value=f"=SUM(I7:I{total_row_idx-1})").number_format = "#,##0.00"
    ws_main.cell(row=total_row_idx, column=9).alignment = Alignment(horizontal="right")

    for col_idx in range(1, len(headers) + 1):
        cell = ws_main.cell(row=total_row_idx, column=col_idx)
        cell.font = total_font
        cell.fill = total_fill
        cell.border = double_bottom_border

    ws_main.row_dimensions[total_row_idx].height = 24

    # Auto-fit column widths
    for col in ws_main.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if cell.row < 5: # Skip titles for width calc
                continue
            if cell.column == 10: # Limit source breakdown column width to avoid extreme width
                max_len = max(max_len, min(len(val_str), 40))
            else:
                max_len = max(max_len, len(val_str))
        ws_main.column_dimensions[col_letter].width = max(max_len + 3, 10)

    # -------------------------------------------------------------------------
    # Sheet 2: DETAIL SOURCES (CHI TIẾT NGUỒN DỮ LIỆU)
    # -------------------------------------------------------------------------
    ws_sources = wb.create_sheet(title="CHI TIET NGUON")
    ws_sources.views.sheetView[0].showGridLines = True

    ws_sources.cell(row=2, column=1, value="CHI TIẾT NGUỒN DỮ LIỆU PHÂN BỔ").font = title_font
    ws_sources.cell(row=3, column=1, value="Danh sách chi tiết các thanh nhôm được trích xuất từ từng file dự án đơn lẻ trước khi gộp.").font = subtitle_font

    source_headers = [
        "STT", "File Nguồn (Project)", "Mã Nhôm (Code)", "Mô Tả (Description)", 
        "Chiều Dài (mm)", "Màu Sắc (Color)", "Số Lượng (Thanh)", 
        "Đơn Trọng (kg/m)", "Khối Lượng (kg)"
    ]

    src_header_row = 5
    for col_idx, h in enumerate(source_headers, 1):
        cell = ws_sources.cell(row=src_header_row, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = PatternFill(start_color="334155", end_color="334155", fill_type="solid") # Slate Header
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        
    ws_sources.row_dimensions[src_header_row].height = 24

    src_curr_row = 6
    for idx, row_data in enumerate(detailed_sources, 1):
        ws_sources.cell(row=src_curr_row, column=1, value=idx).alignment = Alignment(horizontal="center")
        ws_sources.cell(row=src_curr_row, column=2, value=row_data['source_file']).alignment = Alignment(horizontal="left")
        ws_sources.cell(row=src_curr_row, column=3, value=row_data['code']).font = Font(name=font_family, size=11, bold=True)
        ws_sources.cell(row=src_curr_row, column=4, value=row_data['description']).alignment = Alignment(horizontal="left")
        
        l_cell = ws_sources.cell(row=src_curr_row, column=5, value=row_data['length'])
        l_cell.number_format = "#,##0"
        l_cell.alignment = Alignment(horizontal="right")
        
        ws_sources.cell(row=src_curr_row, column=6, value=row_data['color']).alignment = Alignment(horizontal="center")
        
        q_cell = ws_sources.cell(row=src_curr_row, column=7, value=row_data['pieces'])
        q_cell.number_format = "#,##0"
        q_cell.alignment = Alignment(horizontal="right")
        
        w_cell = ws_sources.cell(row=src_curr_row, column=8, value=row_data['unit_weight'])
        w_cell.number_format = "0.000"
        w_cell.alignment = Alignment(horizontal="right")
        
        tw_cell = ws_sources.cell(row=src_curr_row, column=9, value=f"=G{src_curr_row}*E{src_curr_row}/1000*H{src_curr_row}")
        tw_cell.number_format = "#,##0.00"
        tw_cell.alignment = Alignment(horizontal="right")

        for col_idx in range(1, len(source_headers) + 1):
            cell = ws_sources.cell(row=src_curr_row, column=col_idx)
            cell.font = data_font
            cell.border = thin_border
            if idx % 2 == 0:
                cell.fill = zebra_fill
                
        ws_sources.row_dimensions[src_curr_row].height = 20
        src_curr_row += 1

    # Total row for sources sheet
    src_total_row = src_curr_row
    ws_sources.cell(row=src_total_row, column=1, value="Tổng cộng:").alignment = Alignment(horizontal="center")
    ws_sources.cell(row=src_total_row, column=7, value=f"=SUM(G6:G{src_total_row-1})").number_format = "#,##0"
    ws_sources.cell(row=src_total_row, column=7).alignment = Alignment(horizontal="right")
    ws_sources.cell(row=src_total_row, column=9, value=f"=SUM(I6:I{src_total_row-1})").number_format = "#,##0.00"
    ws_sources.cell(row=src_total_row, column=9).alignment = Alignment(horizontal="right")

    for col_idx in range(1, len(source_headers) + 1):
        cell = ws_sources.cell(row=src_total_row, column=col_idx)
        cell.font = total_font
        cell.fill = total_fill
        cell.border = double_bottom_border

    ws_sources.row_dimensions[src_total_row].height = 24

    for col in ws_sources.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if cell.row < 4:
                continue
            max_len = max(max_len, len(val_str))
        ws_sources.column_dimensions[col_letter].width = max(max_len + 3, 10)

    # Save
    wb.save(output_path)
    print(f"Consolidated aluminum orders saved successfully to {output_path}")
    return True, sorted_items


if __name__ == "__main__":
    # Test script locally
    base_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    # 1. Parse opera file
    parse_opera_file(1, os.path.join(base_dir, "5017.xls"))
    # 2. Run calculations
    results = calculate_project_estimates(1)
    print(f"Calculated {len(results)} doors.")
    # 3. Export
    generate_excel_report(1, os.path.join(base_dir, "BAO GIA-NHOM KINH NOVA EC.xlsx"), os.path.join(base_dir, "BAO_GIA_CALCULATED.xlsx"))
